"""
dashboard_vendas/server.py
---------------------------
Servidor local (só biblioteca padrão, sem dependências novas) para o painel
de análise de vendas por produto/família do Mercado Livre.

Uso:
    python dashboard_vendas/server.py
    (abre automaticamente http://localhost:8765)

Aceita, na busca:
  - um MLB ID (anúncio individual) -> busca vendas só desse item
  - um Family ID / group_id (14+ dígitos) -> resolve todas as variações
    (user_products_ids) e, para cada uma, busca vendas em TODOS os item_ids
    associados (o ativo e qualquer fechado/migrado) -- o histórico de venda
    muitas vezes fica no anúncio antigo, não no atual (confirmado: variação
    MLBU3076710630 tinha 0 vendas no item ativo e 113 no fechado).
  - um SKU (código do vendedor) -> resolve via search_by_sku (já existente
    em connectors/mercadolivre/client.py) pra todos os MLB IDs associados.
    Fallback: só é tentado se não bater como MLB nem como Family -- SKUs
    desse projeto costumam ter 14 dígitos, mesmo tamanho de um Family ID.
  - campo vazio -> busca TODOS os produtos vendidos no período selecionado
    (exige um filtro de data específico, não pode ser "Todos" -- o histórico
    completo do vendedor é grande demais pra isso).

Cache: os pedidos são cacheados em disco por dia civil (dashboard_vendas/cache/
AAAA-MM-DD.json), sem filtro de item -- um único cache por dia serve os três
modos de busca (MLB, Family, campo vazio), já que /orders/search sem item= já
devolve todos os pedidos do dia; pra MLB/Family, filtra localmente em memória
em vez de chamar a API de novo com item=. Hoje nunca é lido do cache (sempre
busca ao vivo); dias passados ficam cacheados pra sempre. Ver _obter_pedidos_por_intervalo.

SKU/título por item (cache/skus.json) e tarifa de envio por shipping_id
(cache/fretes.json) também ficam em cache sem expiração -- são fatos
historicamente estáveis (ver _resolver_skus/_resolver_fretes), diferente dos
pedidos de "hoje" que ainda podem mudar.
"""

import io
import json
import re
import sys
import time
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse, parse_qs

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill

from connectors.mercadolivre.client import ml_get, get_item, extract_seller_sku, search_by_sku, ML_USER_ID

_BASE = "https://api.mercadolibre.com"
_PORT = 8765
_HTML_PATH = Path(__file__).parent / "index.html"
_MAX_WORKERS = 8   # concorrência das chamadas ao ML -- uma busca de família (dezenas
                    # de variações x múltiplos item_ids cada) é lenta demais em série
                    # (146s medido pra 64 itens) pro navegador aguentar sem cair a conexão
_MESES_MAX = 3      # limite de meses pra trás permitido no filtro "mês específico"
_CACHE_DIR = Path(__file__).parent / "cache"  # mesmo padrão de services/custo_service.py
_CACHE_SKUS_PATH = _CACHE_DIR / "skus.json"      # {item_id: {"sku":..., "titulo":...}}
_CACHE_FRETES_PATH = _CACHE_DIR / "fretes.json"   # {shipping_id (str): custo}


def _ler_cache_json(caminho: Path) -> dict:
    if not caminho.exists():
        return {}
    try:
        dados = json.loads(caminho.read_text(encoding="utf-8"))
        return dados if isinstance(dados, dict) else {}
    except Exception:
        return {}


def _gravar_cache_json(caminho: Path, dados: dict) -> None:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    tmp = caminho.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(dados, ensure_ascii=False), encoding="utf-8")
    tmp.replace(caminho)  # escrita atômica, mesmo padrão do cache por dia


# ── Filtro de período ──────────────────────────────────────────────────────────
#
# Empurra o filtro pra dentro da própria chamada ML (order.date_created.from/to)
# em vez de buscar tudo e filtrar depois -- pra um item bem vendido isso é a
# diferença entre 1 chamada e várias páginas de pedidos que a gente ia jogar fora.

def _iso_ml(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S.000-03:00")


def _calcular_intervalo(filtro: str, mes: str = ""):
    """Retorna (date_from_iso, date_to_iso) pro filtro pedido, ou (None, None)
    pra 'todos' (sem filtro). Levanta ValueError se o mês específico estiver
    fora da janela permitida (últimos _MESES_MAX meses, incluindo o atual)."""
    agora = datetime.now()

    if filtro == "dia":
        inicio = agora.replace(hour=0, minute=0, second=0, microsecond=0)
        return _iso_ml(inicio), _iso_ml(agora)

    if filtro == "semana":
        inicio = (agora - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        return _iso_ml(inicio), _iso_ml(agora)

    if filtro == "mes":
        inicio = agora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return _iso_ml(inicio), _iso_ml(agora)

    if filtro == "3meses":
        # 3 meses calendário cheios: mês atual + os 2 anteriores, desde o dia 1.
        mes_alvo, ano_alvo = agora.month - 2, agora.year
        while mes_alvo <= 0:
            mes_alvo += 12
            ano_alvo -= 1
        inicio = datetime(ano_alvo, mes_alvo, 1)
        return _iso_ml(inicio), _iso_ml(agora)

    if filtro == "especifico":
        try:
            ano, m = (int(x) for x in mes.split("-"))
            inicio = datetime(ano, m, 1)
        except Exception:
            raise ValueError('Mês inválido -- use o formato "AAAA-MM".')

        meses_atras = (agora.year - ano) * 12 + (agora.month - m)
        if meses_atras < 0 or meses_atras >= _MESES_MAX:
            raise ValueError(f"Escolha um mês dentro dos últimos {_MESES_MAX} meses.")

        fim = datetime(ano + 1, 1, 1) if m == 12 else datetime(ano, m + 1, 1)
        fim -= timedelta(seconds=1)
        return _iso_ml(inicio), _iso_ml(min(fim, agora))

    return None, None  # "todos" (ou valor desconhecido -- sem filtro)


# ── Cache de pedidos por dia civil ──────────────────────────────────────────────
#
# GET /orders/search?seller=X&order.date_created.from=A&to=B (sem item=) já
# devolve TODOS os pedidos do vendedor nessa janela -- um superconjunto que
# contém qualquer pedido que bateria com item=X pra essa mesma janela (o
# filtro item= só restringe se aquele item aparece em order_items[], não muda
# o universo de pedidos -- confirmado testando ao vivo). Por isso cacheamos os
# pedidos BRUTOS por dia civil, sem filtro de item: o mesmo cache serve busca
# por MLB, por Family (várias variações) e por período inteiro (campo vazio)
# -- pra MLB/Family, filtra localmente em memória (_filtra_pedidos_por_item)
# em vez de chamar a API nova de novo com item=.
#
# Hoje nunca é lido do cache (pode ainda estar recebendo pedidos novos) --
# sempre busca ao vivo. Dias passados ficam cacheados pra sempre (imutáveis).

def _dia_str(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def _hoje_str() -> str:
    return _dia_str(datetime.now().date())


def _limites_dia(dia: date) -> tuple:
    inicio = datetime(dia.year, dia.month, dia.day, 0, 0, 0)
    fim = datetime(dia.year, dia.month, dia.day, 23, 59, 59)
    return _iso_ml(inicio), _iso_ml(fim)


def _dias_no_intervalo(date_from_iso: str, date_to_iso: str) -> list:
    # formato fixo de _iso_ml: "AAAA-MM-DDTHH:MM:SS.000-03:00" -- os 10
    # primeiros caracteres já são a data, sem precisar parsear o resto.
    inicio = datetime.strptime(date_from_iso[:10], "%Y-%m-%d").date()
    fim = datetime.strptime(date_to_iso[:10], "%Y-%m-%d").date()
    dias = []
    atual = inicio
    while atual <= fim:
        dias.append(atual)
        atual += timedelta(days=1)
    return dias


def _cache_path_dia(dia_str: str) -> Path:
    return _CACHE_DIR / f"{dia_str}.json"


def _ler_cache_dia(dia_str: str):
    """None = cache miss (arquivo não existe ou está marcado incompleto --
    ex: sobra de um dia que era 'hoje' quando foi gravado)."""
    caminho = _cache_path_dia(dia_str)
    if not caminho.exists():
        return None
    try:
        dados = json.loads(caminho.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(dados, dict) or not dados.get("completo"):
        return None
    return dados.get("pedidos", [])


def _gravar_cache_dia(dia_str: str, pedidos: list, completo: bool) -> None:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    caminho = _cache_path_dia(dia_str)
    tmp = caminho.with_suffix(".json.tmp")
    tmp.write_text(
        json.dumps({"completo": completo, "pedidos": pedidos}, ensure_ascii=False),
        encoding="utf-8",
    )
    tmp.replace(caminho)  # escrita atômica -- evita JSON corrompido em corrida entre threads/requisições


def _buscar_pedidos_dia_api(dia: date) -> list:
    date_from, date_to = _limites_dia(dia)
    pedidos = []
    offset = 0
    while True:
        params = {
            "seller": ML_USER_ID, "limit": 51, "offset": offset,
            "order.date_created.from": date_from, "order.date_created.to": date_to,
        }
        resp = ml_get(f"{_BASE}/orders/search", params)
        if not isinstance(resp, dict):
            break
        resultados = resp.get("results", [])
        if not resultados:
            break
        pedidos.extend(resultados)
        total = (resp.get("paging") or {}).get("total", 0)
        offset += 51
        if offset >= total or offset >= 5000:
            break
        time.sleep(0.05)
    return pedidos


def _obter_pedidos_por_intervalo(date_from_iso: str, date_to_iso: str) -> list:
    """Orquestrador do cache por dia: descobre os gaps, busca só o que falta
    (em paralelo) e devolve todos os pedidos do intervalo, em ordem cronológica."""
    hoje_str = _hoje_str()
    dias = _dias_no_intervalo(date_from_iso, date_to_iso)

    pedidos_por_dia = {}
    pendentes = []
    for dia in dias:
        dia_str = _dia_str(dia)
        if dia_str == hoje_str:
            pendentes.append(dia)
            continue
        cache = _ler_cache_dia(dia_str)
        if cache is not None:
            pedidos_por_dia[dia_str] = cache
        else:
            pendentes.append(dia)

    print(f"[cache] {len(dias) - len(pendentes)}/{len(dias)} dia(s) em cache, buscando {len(pendentes)} na API")

    if pendentes:
        with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as pool:
            futuros = {pool.submit(_buscar_pedidos_dia_api, dia): dia for dia in pendentes}
            for futuro in as_completed(futuros):
                dia = futuros[futuro]
                dia_str = _dia_str(dia)
                pedidos_dia = futuro.result()
                pedidos_por_dia[dia_str] = pedidos_dia
                _gravar_cache_dia(dia_str, pedidos_dia, completo=(dia_str != hoje_str))

    todos = []
    for dia in dias:
        todos.extend(pedidos_por_dia.get(_dia_str(dia), []))
    return todos


def _filtra_pedidos_por_item(pedidos: list, item_id: str) -> list:
    return [
        p for p in pedidos
        if any((oi.get("item") or {}).get("id") == item_id for oi in (p.get("order_items") or []))
    ]


# ── Resolução de entrada (MLB individual x Family ID x SKU) ────────────────────

def _buscar_itens_do_up(up: str) -> list:
    resp = ml_get(f"{_BASE}/users/{ML_USER_ID}/items/search", {"user_product_id": up})
    return [
        {"item_id": item_id, "user_product_id": up}
        for item_id in (resp.get("results", []) if isinstance(resp, dict) else [])
    ]


def _coletar_itens_da_familia(family_id: str) -> list:
    """Retorna [{item_id, user_product_id}] -- TODOS os item_ids de cada
    user_product_id (ativo + fechados/migrados), pra não perder histórico.
    Resolve os user_products_ids em paralelo (uma família com 30+ variações,
    em série, é o principal motivo da busca demorar minutos)."""
    fam = ml_get(f"{_BASE}/sites/MLB/user-products-families/{family_id}")
    ups = fam.get("user_products_ids", []) if isinstance(fam, dict) else []

    itens = []
    with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as pool:
        futuros = [pool.submit(_buscar_itens_do_up, up) for up in ups]
        for futuro in as_completed(futuros):
            itens.extend(futuro.result())
    return itens


def _resolver_tipo_busca(produto_bruto: str):
    """Decide como resolver a entrada em item_ids reais -- MLB direto, Family
    ID, ou SKU (fallback, via search_by_sku já existente em
    connectors/mercadolivre/client.py). Retorna (tipo, itens_info); tipo é
    None e itens_info [] se nada bateu com nenhuma das três formas.

    SKUs desse projeto costumam ter 14 dígitos (mesmo tamanho de um Family
    ID) -- por isso Family só é aceito como tipo final se realmente resolver
    alguma variação; senão, cai pra tentar como SKU mesmo assim."""
    bruto = produto_bruto.strip()

    if bruto.upper().startswith("MLB"):
        return "item", [{"item_id": bruto.upper(), "user_product_id": None}]

    digitos = re.sub(r"\D", "", bruto)
    if len(digitos) >= 14:
        itens_info = _coletar_itens_da_familia(digitos)
        if itens_info:
            return "familia", itens_info

    mlb_ids = search_by_sku(bruto)
    if mlb_ids:
        return "sku", [{"item_id": mid, "user_product_id": None} for mid in mlb_ids]

    return None, []


# ── Pedidos ────────────────────────────────────────────────────────────────────

def _buscar_pedidos_item(item_id: str, date_from: str = None, date_to: str = None,
                          limite_seguranca: int = 5000) -> list:
    pedidos = []
    offset = 0
    while True:
        params = {"seller": ML_USER_ID, "item": item_id, "limit": 51, "offset": offset}
        if date_from:
            params["order.date_created.from"] = date_from
        if date_to:
            params["order.date_created.to"] = date_to
        resp = ml_get(f"{_BASE}/orders/search", params)
        if not isinstance(resp, dict):
            break
        resultados = resp.get("results", [])
        if not resultados:
            break
        pedidos.extend(resultados)
        total = (resp.get("paging") or {}).get("total", 0)
        offset += 51
        if offset >= total or offset >= limite_seguranca:
            break
        time.sleep(0.1)
    return pedidos


# ── Tarifa de envio (Mercado Envios "por sua conta") ───────────────────────────
#
# NÃO vem em /orders/search nem em /orders/{id} -- só a referência shipping.id.
# O valor real cobrado do vendedor está em GET /shipments/{id}/costs, campo
# senders[].cost (confirmado contra a tela de detalhe do pedido do próprio ML:
# "Tarifa do Mercado Envios (por sua conta)"). 1 chamada extra por pedido.

def _buscar_custo_envio(shipping_id) -> float:
    try:
        resp = ml_get(f"{_BASE}/shipments/{shipping_id}/costs")
        senders = resp.get("senders", []) if isinstance(resp, dict) else []
        for s in senders:
            if str(s.get("user_id")) == str(ML_USER_ID):
                return float(s.get("cost", 0) or 0)
        return float(senders[0].get("cost", 0) or 0) if senders else 0.0
    except Exception:
        return 0.0


# Custo de um envio já criado é fato histórico fixo (não muda depois) -- cache
# em disco sem expiração, chaveado por shipping_id. Só busca na API os que
# ainda não apareceram em nenhuma busca anterior.
def _resolver_fretes(shipping_ids) -> dict:
    shipping_ids = [sid for sid in shipping_ids if sid]
    if not shipping_ids:
        return {}

    cache = _ler_cache_json(_CACHE_FRETES_PATH)
    faltantes = [sid for sid in shipping_ids if str(sid) not in cache]

    if faltantes:
        with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as pool:
            futuros = {pool.submit(_buscar_custo_envio, sid): sid for sid in faltantes}
            for futuro in as_completed(futuros):
                sid = futuros[futuro]
                cache[str(sid)] = futuro.result()
        _gravar_cache_json(_CACHE_FRETES_PATH, cache)

    return {sid: cache.get(str(sid), 0.0) for sid in shipping_ids}


# SKU/título de um anúncio quase nunca mudam -- mesma lógica de cache sem
# expiração, chaveado por item_id (MLB).
def _resolver_skus(item_ids) -> dict:
    item_ids = [iid for iid in item_ids if iid]
    if not item_ids:
        return {}

    cache = _ler_cache_json(_CACHE_SKUS_PATH)
    faltantes = [iid for iid in item_ids if iid not in cache]

    if faltantes:
        with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as pool:
            futuros = {pool.submit(get_item, iid): iid for iid in faltantes}
            for futuro in as_completed(futuros):
                iid = futuros[futuro]
                item = futuro.result() or {}
                cache[iid] = {"sku": extract_seller_sku(item), "titulo": item.get("title", "")}
        _gravar_cache_json(_CACHE_SKUS_PATH, cache)

    vazio = {"sku": "", "titulo": ""}
    return {iid: cache.get(iid, vazio) for iid in item_ids}


# "venda total" abaixo desse valor + frete acima desse outro valor = frete
# desproporcional ao tamanho da venda (indício de frete acima do normal pra
# esse produto). Confirmado com um caso real: 3 un. x R$78,99 = R$236,97 de
# venda, R$24,15 de frete -- ok pra esse volume; mas 1 un. sozinha a R$78,99
# com o mesmo frete de R$24,15 já seria desproporcional.
_LIMITE_VENDA_FRETE_ALTO = 78.99
_LIMITE_FRETE_ALTO = 9.0


def _linha_de_item(oi: dict, pedido: dict, tarifa_envio: float) -> dict:
    item = oi.get("item", {}) or {}
    var_attrs = item.get("variation_attributes") or []
    variacao = ", ".join(f"{a.get('name')}: {a.get('value_name')}" for a in var_attrs)

    # unit_price/sale_fee são POR UNIDADE (confirmado: 3un. x 78,99 =
    # 236,97 = total_amount do pedido; 9,36 x 3 = 28,08 = tarifa de 12%
    # mostrada no detalhe do pedido no próprio ML) -- multiplica pela
    # quantidade pra refletir o valor real movimentado nessa linha.
    quantidade = oi.get("quantity", 0) or 0
    venda_total = round((oi.get("unit_price", 0) or 0) * quantidade, 2)
    taxa_total = round((oi.get("sale_fee", 0) or 0) * quantidade, 2)

    return {
        "order_id": pedido.get("id"),
        "pack_id": pedido.get("pack_id"),  # varios pedidos podem compartilhar 1 pacote/envio
        "shipping_id": (pedido.get("shipping") or {}).get("id"),  # chave real pra deduplicar frete na soma
        "data": pedido.get("date_created"),
        "status": pedido.get("status"),
        "fulfilled": bool(pedido.get("fulfilled")),
        "mlb_id": item.get("id"),
        "titulo": item.get("title", ""),
        "variacao": variacao,
        "quantidade": quantidade,
        "preco_unitario": oi.get("unit_price", 0),
        "taxa_ml": oi.get("sale_fee", 0),
        "venda_total": venda_total,
        "taxa_total": taxa_total,
        "tarifa_envio": tarifa_envio,
        "frete_alto": venda_total < _LIMITE_VENDA_FRETE_ALTO and tarifa_envio > _LIMITE_FRETE_ALTO,
        "listing_type_id": oi.get("listing_type_id", ""),
        "total_pedido": pedido.get("total_amount", 0),
    }


def _linhas_de_pedido(pedido: dict, item_id_alvo: str, custo_envio_por_shipping: dict) -> list:
    shipping_id = (pedido.get("shipping") or {}).get("id")
    tarifa_envio = custo_envio_por_shipping.get(shipping_id, 0.0) if shipping_id else 0.0
    return [
        _linha_de_item(oi, pedido, tarifa_envio)
        for oi in (pedido.get("order_items", []) or [])
        if (oi.get("item") or {}).get("id") == item_id_alvo
    ]


# ── Busca por período, sem produto específico (todos os produtos) ─────────────
#
# Sem MLB/Family ID pra resolver primeiro -- busca os PEDIDOS direto por data
# (via _obter_pedidos_por_intervalo, cache por dia) e só depois descobre quais
# produtos apareceram, resolvendo SKU/frete pra cada um. "Todos" fica de fora
# por segurança: o vendedor já tem 34 mil+ pedidos pagos no total, buscar tudo
# sem filtro de período travaria a busca.

def buscar_vendas_todos_produtos(filtro: str = "todos", mes: str = "") -> dict:
    date_from, date_to = _calcular_intervalo(filtro, mes)  # pode levantar ValueError
    if filtro == "todos":
        raise ValueError(
            'Pra buscar todos os produtos, escolha um período (não pode ser "Todos") -- '
            "o histórico completo do vendedor é grande demais pra isso."
        )

    pedidos = _obter_pedidos_por_intervalo(date_from, date_to)
    if not pedidos:
        return {
            "ok": True, "produto_buscado": "(todos os produtos)", "tipo": "todos",
            "itens_consultados": [], "vendas": [],
            "filtro_aplicado": {"filtro": filtro, "date_from": date_from, "date_to": date_to},
        }

    shipping_ids = {
        (p.get("shipping") or {}).get("id") for p in pedidos if (p.get("shipping") or {}).get("id")
    }
    custo_envio_por_shipping = _resolver_fretes(shipping_ids)

    item_ids = {
        (oi.get("item") or {}).get("id")
        for p in pedidos for oi in (p.get("order_items") or [])
        if (oi.get("item") or {}).get("id")
    }
    sku_por_item = _resolver_skus(item_ids)

    vendas = []
    for pedido in pedidos:
        shipping_id = (pedido.get("shipping") or {}).get("id")
        tarifa_envio = custo_envio_por_shipping.get(shipping_id, 0.0) if shipping_id else 0.0
        for oi in pedido.get("order_items", []) or []:
            linha = _linha_de_item(oi, pedido, tarifa_envio)
            linha["sku"] = sku_por_item.get(linha["mlb_id"], {}).get("sku", "")
            linha["produto_buscado"] = "(todos os produtos)"
            vendas.append(linha)

    return {
        "ok": True,
        "produto_buscado": "(todos os produtos)",
        "tipo": "todos",
        "itens_consultados": sorted(item_ids),
        "vendas": vendas,
        "filtro_aplicado": {"filtro": filtro, "date_from": date_from, "date_to": date_to},
    }


# ── Handler principal da busca ─────────────────────────────────────────────────

def buscar_vendas_produto(produto_bruto: str, filtro: str = "todos", mes: str = "") -> dict:
    date_from, date_to = _calcular_intervalo(filtro, mes)  # pode levantar ValueError

    tipo, itens_info = _resolver_tipo_busca(produto_bruto)
    if not itens_info:
        return {"ok": False, "erro": f'Não encontrei nada pra "{produto_bruto}" (tentei como MLB ID, Family ID e SKU).'}

    item_ids = [i["item_id"] for i in itens_info]

    info_por_item = _resolver_skus(item_ids)

    # Com período delimitado, busca os pedidos UMA VEZ só (compartilhado, via
    # cache por dia) e filtra localmente por item -- antes, cada item_id da
    # família disparava sua própria busca na API; agora N variações == 1 busca
    # de pedidos + N filtros em memória. Sem período ("todos"), sem cache: cai
    # no caminho antigo, 1 busca por item direto na API (mesmo de sempre).
    if date_from and date_to:
        pedidos_compartilhados = _obter_pedidos_por_intervalo(date_from, date_to)
        pedidos_por_item = {
            item_id: _filtra_pedidos_por_item(pedidos_compartilhados, item_id)
            for item_id in item_ids
        }
    else:
        pedidos_por_item = {}
        with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as pool:
            futuros = {
                pool.submit(_buscar_pedidos_item, item_id, date_from, date_to): item_id
                for item_id in item_ids
            }
            for futuro in as_completed(futuros):
                item_id = futuros[futuro]
                pedidos_por_item[item_id] = futuro.result()

    # Tarifa de envio ("por sua conta") -- cacheada por shipping_id (fato
    # histórico fixo, ver _resolver_fretes), só busca na API o que faltar.
    shipping_ids = {
        (pedido.get("shipping") or {}).get("id")
        for pedidos in pedidos_por_item.values()
        for pedido in pedidos
        if (pedido.get("shipping") or {}).get("id")
    }
    custo_envio_por_shipping = _resolver_fretes(shipping_ids)

    vendas = []
    for item_id in item_ids:
        for pedido in pedidos_por_item.get(item_id, []):
            for linha in _linhas_de_pedido(pedido, item_id, custo_envio_por_shipping):
                linha["sku"] = info_por_item.get(item_id, {}).get("sku", "")
                if not linha["titulo"]:
                    linha["titulo"] = info_por_item.get(item_id, {}).get("titulo", "")
                linha["produto_buscado"] = produto_bruto
                vendas.append(linha)

    return {
        "ok": True,
        "produto_buscado": produto_bruto,
        "tipo": tipo,
        "itens_consultados": [i["item_id"] for i in itens_info],
        "vendas": vendas,
        "filtro_aplicado": {"filtro": filtro, "date_from": date_from, "date_to": date_to},
    }


# ── Exportar Excel ──────────────────────────────────────────────────────────────

_COLUNAS_EXPORT = [
    ("Data",              "data"),
    ("Pedido",            "order_id"),
    ("Pacote",            "pack_id"),
    ("Busca de origem",   "produto_buscado"),
    ("MLB",               "mlb_id"),
    ("SKU",               "sku"),
    ("Variação",          "variacao"),
    ("Título",            "titulo"),
    ("Qtd",               "quantidade"),
    ("Venda Total (R$)",  "venda_total"),
    ("Taxa ML (R$)",      "taxa_total"),
    ("Tarifa Envio (R$)", "tarifa_envio"),
    ("Frete Alto?",       "frete_alto"),
    ("Tipo Anúncio",      "listing_type_id"),
    ("Status",            "status"),
    ("Full",              "fulfilled"),
]
_CHAVES_MOEDA = {"venda_total", "taxa_total", "tarifa_envio"}
_COR_FRETE_ALTO = "FCE4CC"  # laranja claro -- mesma leitura visual da tabela web
_ROTULO_TIPO_ANUNCIO = {"gold_special": "Clássico", "gold_pro": "Premium"}


def gerar_xlsx(vendas: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"

    ws.append([titulo for titulo, _ in _COLUNAS_EXPORT])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for v in vendas:
        linha = []
        for _, chave in _COLUNAS_EXPORT:
            valor = v.get(chave, "")
            if chave in ("fulfilled", "frete_alto"):
                valor = "Sim" if valor else "Não"
            elif chave == "listing_type_id":
                valor = _ROTULO_TIPO_ANUNCIO.get(valor, valor)
            linha.append(valor)
        ws.append(linha)

    idx_moeda = [i for i, (_, k) in enumerate(_COLUNAS_EXPORT) if k in _CHAVES_MOEDA]
    preenchimento = PatternFill(start_color=_COR_FRETE_ALTO, end_color=_COR_FRETE_ALTO, fill_type="solid")
    for i, v in enumerate(vendas):
        row = ws[i + 2]
        for idx in idx_moeda:
            row[idx].number_format = '"R$"#,##0.00'
        if v.get("frete_alto"):
            for cell in row:
                cell.fill = preenchimento

    for col_cells in ws.columns:
        largura = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(largura + 2, 45)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Servidor HTTP ──────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass  # silencia o log padrão no console

    def _json(self, payload: dict, status: int = 200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/":
            html = _HTML_PATH.read_text(encoding="utf-8")
            body = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/vendas":
            qs = parse_qs(parsed.query)
            produto = (qs.get("produto") or [""])[0].strip()
            filtro = (qs.get("filtro") or ["todos"])[0].strip()
            mes = (qs.get("mes") or [""])[0].strip()
            try:
                if produto:
                    resultado = buscar_vendas_produto(produto, filtro, mes)
                else:
                    # Sem MLB/Family ID -- busca todos os produtos vendidos no período.
                    resultado = buscar_vendas_todos_produtos(filtro, mes)
                self._json(resultado)
            except ValueError as e:
                self._json({"ok": False, "erro": str(e)}, 400)
            except Exception as e:
                self._json({"ok": False, "erro": str(e)}, 500)
            return

        self.send_response(404)
        self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path == "/api/exportar":
            length = int(self.headers.get("Content-Length", 0) or 0)
            corpo = self.rfile.read(length) if length else b""
            try:
                vendas = json.loads(corpo.decode("utf-8")) if corpo else []
                if not isinstance(vendas, list):
                    raise ValueError("formato inválido")
                xlsx_bytes = gerar_xlsx(vendas)
            except Exception as e:
                self._json({"ok": False, "erro": f"Erro ao gerar Excel: {e}"}, 400)
                return

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="vendas_bouwobra.xlsx"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            return

        self.send_response(404)
        self.end_headers()


def main():
    server = ThreadingHTTPServer(("localhost", _PORT), Handler)
    url = f"http://localhost:{_PORT}"
    print(f"Dashboard de Vendas rodando em {url}  (Ctrl+C pra parar)")
    try:
        webbrowser.open(url)
    except Exception:
        pass
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nEncerrado.")


if __name__ == "__main__":
    main()

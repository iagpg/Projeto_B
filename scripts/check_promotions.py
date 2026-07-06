"""
scripts/check_promotions.py
----------------------------
Verifica promoções ML para anúncios de uma aba da Curva do Ecommerce.xlsx
e cria uma aba "Analise Promocoes <CURVA>" com o status de cada campanha.

Campanhas monitoradas:
  C-MLB4305989  "Junho ate Julho"  — expira 01/07/2026  ⚠ URGENTE
  C-MLB4586868  "Julho"            — inicia 01/07/2026, vai até 01/08/2026

Tipos de linha suportados na coluna A:
  - IDs individuais (7-13 dígitos) → MLB{id}
  - Group IDs (14-16 dígitos)      → MLBU{id}, expande todas as variações

Linhas com "Estoque Encalhado" na coluna R são ignoradas.

Uso:
    python scripts/check_promotions.py        # padrão: Curva B
    python scripts/check_promotions.py --a    # Curva A
    python scripts/check_promotions.py --b    # Curva B
    python scripts/check_promotions.py --c    # Curva C
"""

import argparse
import json
import sys
import time
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

ROOT        = Path(__file__).parent.parent
cfg         = json.load(open(ROOT / "config.json", encoding="utf-8"))
USER_ID     = cfg["ml_user_id"]
EXCEL_PATH  = ROOT / "output" / "Ecommerce.xlsx"
BASE        = "https://api.mercadolibre.com"

CAMP_JUNHO  = "C-MLB4305989"   # Junho ate Julho — expira 01/07
CAMP_JULHO  = "C-MLB4586868"   # Julho           — inicia 01/07

COL_ID      = 1    # Coluna A: IDs dos anúncios
COL_STATUS  = 18   # Coluna R: classificação (ex: "Estoque Encalhado")

_token = cfg.get("ml_access_token", "")


# ── Token ─────────────────────────────────────────────────────────────────────

def _refresh():
    global _token, cfg
    r = requests.post(f"{BASE}/oauth/token", data={
        "grant_type":    "refresh_token",
        "client_id":     cfg["ml_client_id"],
        "client_secret": cfg["ml_client_secret"],
        "refresh_token": cfg["ml_refresh_token"],
    }, timeout=15)
    d = r.json()
    if "access_token" not in d:
        sys.exit(f"Token refresh failed: {d}")
    _token = d["access_token"]
    cfg["ml_access_token"]  = _token
    cfg["ml_refresh_token"] = d.get("refresh_token", cfg["ml_refresh_token"])
    with open(ROOT / "config.json", "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)


def _get(url, params=None, _retry=True):
    h = {"Authorization": f"Bearer {_token}"}
    r = requests.get(url, headers=h, params=params or {}, timeout=15)
    if r.status_code == 401 and _retry:
        _refresh()
        return _get(url, params, _retry=False)
    if r.status_code == 429:
        time.sleep(3)
        return _get(url, params, _retry=_retry)
    if r.status_code in (400, 404):
        return None
    if not r.ok:
        return None
    return r.json()


# ── ML helpers ────────────────────────────────────────────────────────────────

def get_item_details_batch(mlb_ids):
    """Busca título, preço e family_id de até 20 itens em uma chamada."""
    if not mlb_ids:
        return {}
    data = _get(f"{BASE}/items", {
        "ids":        ",".join(mlb_ids[:20]),
        "attributes": "id,title,price,status,family_id",
    })
    result = {}
    if isinstance(data, list):
        for entry in data:
            body = entry.get("body") if isinstance(entry, dict) else entry
            if body and body.get("id"):
                result[body["id"]] = body
    return result


def get_item_promotions(mlb_id):
    """Retorna lista de promoções do item."""
    data = _get(f"{BASE}/seller-promotions/items/{mlb_id}?app_version=v2")
    return data if isinstance(data, list) else []


def build_family_map() -> dict:
    """
    Varre TODOS os itens do vendedor (paginado) e mapeia family_id → [MLB IDs].
    Chamado uma única vez quando existem group IDs a resolver.
    O parâmetro family_id no endpoint de busca é ignorado pela API ML;
    a única forma confiável é buscar o campo family_id item a item.
    """
    print("  Mapeando familias dos itens do vendedor (necessario para grupos)...")
    all_ids = []
    offset  = 0
    limit   = 50
    while True:
        data = _get(f"{BASE}/users/{USER_ID}/items/search",
                    {"limit": limit, "offset": offset})
        if not isinstance(data, dict):
            break
        results = data.get("results", [])
        all_ids.extend(results)
        total = data.get("paging", {}).get("total", 0)
        offset += limit
        if offset >= total or not results:
            break
        time.sleep(0.1)
    print(f"  Total de itens do vendedor: {len(all_ids)}")

    family_map = {}   # family_id (str) → [MLB IDs]
    item_family = {}  # MLB ID → family_id  (usado para validação posterior)
    for i in range(0, len(all_ids), 20):
        chunk = all_ids[i:i+20]
        data = _get(f"{BASE}/items", {
            "ids":        ",".join(chunk),
            "attributes": "id,family_id,title,status",
        })
        if isinstance(data, list):
            for entry in data:
                body = entry.get("body", {}) if isinstance(entry, dict) else {}
                fid  = str(body.get("family_id") or "").strip()
                mlb  = body.get("id", "")
                if fid and mlb:
                    family_map.setdefault(fid, []).append(mlb)
                    item_family[mlb] = fid
        time.sleep(0.1)

    total_families = len(family_map)
    total_mapped   = sum(len(v) for v in family_map.values())
    print(f"  Familias encontradas: {total_families} | Itens mapeados: {total_mapped}")
    return family_map


def parse_cell_ids(raw_val):
    """
    Extrai IDs de uma célula. Retorna (item_ids, group_ids).
      item_ids:  dígitos 7-13 (puro ou prefixo MLB)  → anúncios individuais
      group_ids: dígitos 14+  (puro ou prefixo MLBU) → grupos de variações

    Aceita formatos com ou sem prefixo:
      - "3943414937"          → item
      - "MLB3943414937"       → item
      - "8394429844084229"    → grupo
      - "MLBU8394429844084229"→ grupo (prefixo stripped automaticamente)
    """
    if raw_val is None:
        return [], []
    raw = str(raw_val).strip()
    item_ids  = []
    group_ids = []
    for p in raw.split():
        p = p.strip().upper()
        # MLBU prefix → grupo
        if p.startswith("MLBU"):
            numeric = p[4:]
            if numeric.isdigit() and len(numeric) >= 14:
                group_ids.append(numeric)
            continue
        # MLB prefix → item individual
        if p.startswith("MLB"):
            numeric = p[3:]
            if numeric.isdigit() and 7 <= len(numeric) <= 13:
                item_ids.append(numeric)
            continue
        # Puramente numérico
        if not p.isdigit():
            continue
        n = len(p)
        if 7 <= n <= 13:
            item_ids.append(p)
        elif n >= 14:
            group_ids.append(p)
    return item_ids, group_ids


def check_campaign(promotions, camp_id):
    """Verifica se o item está em uma campanha específica.
    Retorna (in_campaign, status, price)."""
    for promo in promotions:
        if promo.get("id") == camp_id:
            return True, promo.get("status", ""), promo.get("price") or None
    return False, None, None


def camp_text(in_camp, camp_status, camp_price):
    if not in_camp:
        return "Fora"
    if camp_status == "started":
        return f"Ativa (R$ {camp_price:.2f})" if camp_price else "Ativa"
    if camp_status == "pending":
        return f"Inscrito (R$ {camp_price:.2f})" if camp_price else "Inscrito"
    if camp_status == "candidate":
        return "Candidato"
    return camp_status or "—"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Verifica promoções ML por Curva")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--a", dest="curva", action="store_const", const="A", help="Curva A")
    group.add_argument("--b", dest="curva", action="store_const", const="B", help="Curva B (padrão)")
    group.add_argument("--c", dest="curva", action="store_const", const="C", help="Curva C")
    args = parser.parse_args()

    curva    = args.curva or "B"
    tab_name = f"Curva {curva}"
    out_tab  = f"Analise Promocoes {curva}"

    _refresh()
    print(f"Token refreshed OK | Processando: {tab_name}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    if tab_name not in wb.sheetnames:
        sys.exit(f"Aba '{tab_name}' nao encontrada em {EXCEL_PATH}")
    ws = wb[tab_name]
    print(f"{tab_name}: {ws.max_row - 1} linhas de dados")

    # ── Coleta linhas ─────────────────────────────────────────────────────────
    # rows_data: [(row_num, item_ids: [str], group_ids: [str])]
    # item_ids  = dígitos sem prefixo (MLB será adicionado)
    # group_ids = dígitos sem prefixo (MLBU será adicionado)

    rows_data     = []
    skipped       = 0
    all_item_nums = []  # dígitos de IDs individuais

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=COL_STATUS):
        cell_a = row[COL_ID - 1]      # coluna A
        cell_r = row[COL_STATUS - 1]  # coluna R

        r_val = str(cell_r.value or "").strip()
        if "Estoque Encalhado" in r_val:
            skipped += 1
            continue

        item_ids, group_ids = parse_cell_ids(cell_a.value)
        if not item_ids and not group_ids:
            continue

        rows_data.append((cell_a.row, item_ids, group_ids))
        all_item_nums.extend(item_ids)

    print(f"Linhas validas: {len(rows_data)} | Ignoradas (Encalhado): {skipped}")

    # ── Expande grupos → IDs de variações via family_id ──────────────────────

    group_cache   = {}  # gid (numeric str) → [MLB..., ...]
    unique_groups = list(dict.fromkeys(gid for _, _, gids in rows_data for gid in gids))
    all_group_mlbs = []  # MLB IDs vindos de grupos

    family_map = {}
    if unique_groups:
        family_map = build_family_map()

    for gid in unique_groups:
        members = family_map.get(gid, [])
        group_cache[gid] = members
        all_group_mlbs.extend(members)
        status_str = f"{len(members)} variacoes" if members else "NAO ENCONTRADO — verificar family_id"
        print(f"  Grupo {gid}: {status_str}")

    # Todos os MLB IDs únicos (individuais + membros de grupo)
    all_mlb_unique = list(dict.fromkeys(
        [f"MLB{n}" for n in all_item_nums] + all_group_mlbs
    ))
    print(f"IDs unicos totais: {len(all_mlb_unique)}")

    # ── Batch fetch de detalhes (título / preço) ──────────────────────────────

    details_cache = {}
    for i in range(0, len(all_mlb_unique), 20):
        batch = get_item_details_batch(all_mlb_unique[i:i+20])
        details_cache.update(batch)
        time.sleep(0.2)
    print(f"Detalhes obtidos: {len(details_cache)}")

    # Valida membros do grupo usando o family_id retornado no detalhe do item.
    # Como o family_map foi construído com o mesmo atributo, a consistência
    # já está garantida — mas verificamos se o detalhe confirmou o family_id.
    for gid in unique_groups:
        raw_list  = group_cache.get(gid, [])
        validated = [
            mlb_id for mlb_id in raw_list
            if str(details_cache.get(mlb_id, {}).get("family_id") or "").strip() == gid
        ]
        descartados = len(raw_list) - len(validated)
        if descartados:
            print(f"  [AVISO] grupo {gid}: {descartados} itens descartados (family_id divergente)")
        group_cache[gid] = validated
        print(f"  Grupo {gid}: {len(validated)} variacoes confirmadas")

    # ── Fetch de promoções por item ───────────────────────────────────────────

    promotions_cache = {}  # MLB... → [promo, ...]
    for idx, mlb_id in enumerate(all_mlb_unique):
        print(f"  [{idx+1}/{len(all_mlb_unique)}] {mlb_id} ...", end=" ", flush=True)
        promos = get_item_promotions(mlb_id)
        promotions_cache[mlb_id] = promos
        flags = []
        in_j, _, _   = check_campaign(promos, CAMP_JUNHO)
        in_jul, _, _ = check_campaign(promos, CAMP_JULHO)
        if in_j:   flags.append("JUNHO")
        if in_jul: flags.append("JULHO")
        print(", ".join(flags) if flags else "sem campanha")
        time.sleep(0.2)

    # ── Cria nova aba ─────────────────────────────────────────────────────────

    if out_tab in wb.sheetnames:
        del wb[out_tab]
    ws_out = wb.create_sheet(out_tab)

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    RED_FILL    = PatternFill("solid", fgColor="FF4444")
    ORANGE_FILL = PatternFill("solid", fgColor="FFB347")
    GREEN_FILL  = PatternFill("solid", fgColor="92D050")
    YELLOW_FILL = PatternFill("solid", fgColor="FFFF99")
    GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
    BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")  # fundo de linhas de grupo

    headers = [
        "MLB ID / Grupo",
        "Título",
        "Preço Atual (R$)",
        "Status ML",
        "Camp. Junho\n(expira 01/07 ⚠)",
        "Camp. Julho\n(01/07 a 01/08)",
        "Alerta",
    ]

    for col, h in enumerate(headers, 1):
        c = ws_out.cell(row=1, column=col, value=h)
        c.fill      = HEADER_FILL
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    ws_out.row_dimensions[1].height = 36
    for col, w in enumerate([18, 52, 16, 12, 20, 20, 48], 1):
        ws_out.column_dimensions[get_column_letter(col)].width = w

    # ── Helpers de escrita ────────────────────────────────────────────────────

    out_row     = 2
    seen_items  = set()
    seen_groups = set()

    def _junho_fill(in_junho, junho_status, julho_confirmado):
        if in_junho and junho_status == "started" and not julho_confirmado:
            return ORANGE_FILL
        if in_junho:
            return YELLOW_FILL
        return GRAY_FILL

    def _julho_fill(in_julho, julho_status, julho_confirmado):
        if julho_confirmado:
            return GREEN_FILL
        if in_julho and julho_status == "candidate":
            return YELLOW_FILL
        return RED_FILL

    def _alerta_style(alerta_text, alerta_urgente):
        if alerta_urgente:
            return ORANGE_FILL, Font(bold=True, color="880000")
        if "Fora de ambas" in alerta_text or "Verificar" in alerta_text:
            return RED_FILL, Font(color="880000")
        if "Candidato" in alerta_text:
            return YELLOW_FILL, Font(color="664400")
        return GREEN_FILL, Font(color="1A5C1A")

    def write_item_row(mlb_id):
        nonlocal out_row
        if mlb_id in seen_items:
            return
        seen_items.add(mlb_id)

        detail    = details_cache.get(mlb_id, {})
        title     = detail.get("title", "— não encontrado —")
        price     = detail.get("price")
        status_ml = detail.get("status", "—")

        promos = promotions_cache.get(mlb_id, [])
        in_junho, junho_status, junho_price = check_campaign(promos, CAMP_JUNHO)
        in_julho, julho_status, julho_price = check_campaign(promos, CAMP_JULHO)
        julho_confirmado = in_julho and julho_status in ("started", "pending")

        junho_t = camp_text(in_junho, junho_status, junho_price)
        julho_t = camp_text(in_julho, julho_status, julho_price)

        if julho_confirmado and in_junho and junho_status == "started":
            alerta_text    = "Julho ja programado — OK"
            alerta_urgente = False
        elif julho_confirmado:
            alerta_text    = "Julho ativo — OK"
            alerta_urgente = False
        elif in_julho and julho_status == "candidate":
            alerta_text    = "Candidato ao Julho — confirmar inscricao"
            alerta_urgente = False
        elif in_junho and junho_status == "started":
            alerta_text    = "EXPIRA AMANHA! Nao inscrito no Julho"
            alerta_urgente = True
        elif not in_junho and not in_julho:
            alerta_text    = "Fora de ambas as campanhas"
            alerta_urgente = False
        else:
            alerta_text    = "Verificar"
            alerta_urgente = False

        j6_fill  = _junho_fill(in_junho, junho_status, julho_confirmado)
        j7_fill  = _julho_fill(in_julho, julho_status, julho_confirmado)
        a_fill, a_font = _alerta_style(alerta_text, alerta_urgente)

        values = [mlb_id, title, price, status_ml, junho_t, julho_t, alerta_text]
        for col, val in enumerate(values, 1):
            c = ws_out.cell(row=out_row, column=col, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (2, 7)))
            if col == 1:
                c.font = Font(bold=True)
            elif col == 5:
                c.fill = j6_fill
            elif col == 6:
                c.fill = j7_fill
            elif col == 7:
                c.fill = a_fill
                c.font = a_font

        out_row += 1

    def write_group_row(gid):
        nonlocal out_row
        if gid in seen_groups:
            return
        seen_groups.add(gid)

        mlbu_id = f"MLBU{gid}"
        members = group_cache.get(gid, [])
        total   = len(members)

        if total == 0:
            c = ws_out.cell(row=out_row, column=1, value=gid)
            c.font = Font(bold=True, italic=True)
            ws_out.cell(row=out_row, column=2, value="— grupo sem variacoes validadas —")
            ws_out.cell(row=out_row, column=7, value="Grupo vazio / erro de validacao")
            for col in range(1, 8):
                ws_out.cell(row=out_row, column=col).fill = GRAY_FILL
            out_row += 1
            return

        # Agrega status de todas as variações
        julho_ok   = 0
        candidatos = 0
        junho_only = 0
        fora_ambas = 0
        has_junho  = 0
        first_title = None

        for mlb_id in members:
            promos = promotions_cache.get(mlb_id, [])
            in_j,  j_st,  _ = check_campaign(promos, CAMP_JUNHO)
            in_jul, jul_st, _ = check_campaign(promos, CAMP_JULHO)
            jul_ok = in_jul and jul_st in ("started", "pending")

            if in_j:
                has_junho += 1
            if jul_ok:
                julho_ok += 1
            elif in_jul and jul_st == "candidate":
                candidatos += 1
            elif in_j and j_st == "started":
                junho_only += 1
            else:
                fora_ambas += 1

            if first_title is None:
                first_title = details_cache.get(mlb_id, {}).get("title", "")

        base_title = first_title or mlbu_id
        if len(base_title) > 38:
            base_title = base_title[:38] + "…"
        display_title = f"{base_title} (grupo, {total} var.)"

        junho_resumo = f"{has_junho}/{total} em Junho"
        julho_resumo = f"{julho_ok}/{total} com Julho"

        if julho_ok == total:
            alerta_text    = f"Todas as {total} variacoes com Julho OK"
            alerta_urgente = False
            j7_fill        = GREEN_FILL
            a_fill         = GREEN_FILL
            a_font         = Font(bold=True, color="1A5C1A")
        elif julho_ok > 0:
            faltam         = total - julho_ok
            alerta_text    = f"{faltam}/{total} variacoes SEM Julho"
            alerta_urgente = True
            j7_fill        = ORANGE_FILL
            a_fill         = ORANGE_FILL
            a_font         = Font(bold=True, color="880000")
        elif candidatos > 0:
            alerta_text    = f"{candidatos}/{total} candidatas — confirmar inscricao"
            alerta_urgente = False
            j7_fill        = YELLOW_FILL
            a_fill         = YELLOW_FILL
            a_font         = Font(color="664400")
        elif junho_only > 0:
            alerta_text    = f"URGENTE! {junho_only}/{total} apenas em Junho"
            alerta_urgente = True
            j7_fill        = RED_FILL
            a_fill         = ORANGE_FILL
            a_font         = Font(bold=True, color="880000")
        else:
            alerta_text    = f"Nenhuma variacao em campanha"
            alerta_urgente = False
            j7_fill        = RED_FILL
            a_fill         = RED_FILL
            a_font         = Font(color="880000")

        j6_fill = (ORANGE_FILL if junho_only > 0 and julho_ok < total
                   else YELLOW_FILL if has_junho > 0
                   else GRAY_FILL)

        values = [gid, display_title, None, "grupo", junho_resumo, julho_resumo, alerta_text]
        for col, val in enumerate(values, 1):
            c = ws_out.cell(row=out_row, column=col, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (2, 7)))
            c.fill = BLUE_FILL  # fundo padrão azul claro para linhas de grupo
            if col == 1:
                c.font = Font(bold=True, italic=True)
            elif col == 5:
                c.fill = j6_fill
            elif col == 6:
                c.fill = j7_fill
            elif col == 7:
                c.fill = a_fill
                c.font = a_font

        out_row += 1

    # ── Escreve todas as linhas na ordem da planilha ──────────────────────────

    for _, item_ids, group_ids in rows_data:
        for raw_id in item_ids:
            write_item_row(f"MLB{raw_id}")
        for gid in group_ids:
            write_group_row(gid)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{out_row - 1}"

    wb.save(EXCEL_PATH)
    print(f"\nAba '{out_tab}' criada com {out_row - 2} linhas.")
    print(f"Arquivo salvo: {EXCEL_PATH}")

    # ── Resumo ────────────────────────────────────────────────────────────────

    def _has(promos, camp_id, statuses=None):
        return any(p.get("id") == camp_id and (statuses is None or p.get("status") in statuses)
                   for p in promos)

    total_items    = len(all_mlb_unique)
    julho_ok_cnt   = sum(1 for p in promotions_cache.values() if _has(p, CAMP_JULHO, ("started", "pending")))
    junho_only_cnt = sum(1 for p in promotions_cache.values()
                         if _has(p, CAMP_JUNHO, ("started",)) and not _has(p, CAMP_JULHO, ("started", "pending")))
    fora_cnt       = sum(1 for p in promotions_cache.values() if not _has(p, CAMP_JUNHO) and not _has(p, CAMP_JULHO))

    print(f"\n{'='*54}")
    print(f"  Anuncios analisados (incl. variacoes) : {total_items}")
    print(f"  Linhas na aba gerada                  : {out_row - 2}")
    print(f"  Grupos expandidos                     : {len(unique_groups)}")
    print(f"  Ignorados (Estoque Encalhado)          : {skipped}")
    print(f"  Julho ja garantido (OK)               : {julho_ok_cnt}")
    print(f"  Apenas Junho (URGENTE!)               : {junho_only_cnt}")
    print(f"  Fora de ambas campanhas               : {fora_cnt}")
    print(f"{'='*54}")


if __name__ == "__main__":
    main()

"""
scripts/exportar_cache_nf_excel.py
------------------------------------
Exporta o cache/nf_custo.json para um Excel legivel em output/cache_nf.xlsx.

Uso:
    python scripts/exportar_cache_nf_excel.py
"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from services.custo_service import load_cache, cache_info

OUTPUT_DIR  = ROOT / "output"
OUTPUT_FILE = OUTPUT_DIR / "cache_nf.xlsx"

HEADERS = [
    "SKU", "Descricao", "Custo Base (R$)", "IPI (R$)", "Custo c/ IPI (R$)",
    "ICMS Credito (R$)", "PIS Credito (R$)", "COFINS Credito (R$)",
    "Imposto Recuperavel (R$)", "NF Numero", "NF Data", "NF ID",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
FMT_BRL = '"R$"#,##0.00'


def main():
    print("\n" + "=" * 60)
    print("  BouwObra - Exportar Cache NF para Excel")
    print("=" * 60)

    info = cache_info()
    if not info["exists"]:
        print("\n  ERRO: cache/nf_custo.json nao existe.")
        print("  Execute: python scripts/sincronizar_custo.py --full")
        return

    cache = load_cache()
    print(f"\n  Cache: {len(cache)} SKUs (versao: {info['version']})")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cache NF"

    # Cabecalho
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    # Dados, ordenados por SKU
    for sku in sorted(cache.keys(), key=lambda s: s.lower()):
        d = cache[sku]
        ws.append([
            sku,
            d.descricao,
            d.custo_base,
            d.ipi_valor,
            d.custo_com_ipi,
            d.icms_credito,
            d.pis_credito,
            d.cofins_credito,
            d.imposto_recuperavel,
            d.nf_numero,
            d.nf_data,
            d.nf_id,
        ])

    n_rows = ws.max_row

    # Formatos numericos (colunas C..I = 3..9)
    for col in range(3, 10):
        for row in range(2, n_rows + 1):
            ws.cell(row=row, column=col).number_format = FMT_BRL

    # Larguras de coluna
    widths = [16, 42, 14, 12, 15, 14, 13, 15, 16, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{n_rows}"

    wb.save(OUTPUT_FILE)
    print(f"\n  OK: Excel gerado em {OUTPUT_FILE}")
    print(f"  {n_rows - 1} produtos exportados.\n")


if __name__ == "__main__":
    main()

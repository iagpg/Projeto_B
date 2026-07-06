"""
validate_skus.py
----------------
Full pipeline: loads an Excel file, checks each SKU against
Mercado Livre and Tiny ERP APIs, fills Column E, saves output.

Requirements:
    pip install requests openpyxl

Usage:
    python validate_skus.py
"""

import json
import time
import sys
from pathlib import Path
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
sys.path.insert(0, str(Path(__file__).parent.parent))
from connectors.tiny import client as tc

# ── Config ────────────────────────────────────────────────────────────────────

CONFIG_FILE = Path(__file__).parent.parent / "config.json"

with open(CONFIG_FILE) as f:
    cfg = json.load(f)

ML_CLIENT_ID     = cfg["ml_client_id"]
ML_CLIENT_SECRET = cfg["ml_client_secret"]
ML_USER_ID       = cfg["ml_user_id"]
EXCEL_INPUT      = cfg["excel_input"]
EXCEL_OUTPUT     = cfg["excel_output"]

ml_access_token  = cfg.get("ml_access_token", "")
ml_refresh_token = cfg["ml_refresh_token"]

# Mapa family_id → [MLB IDs], populado uma única vez em main() quando necessário
_family_map: dict = {}


# ── ML helpers ────────────────────────────────────────────────────────────────

def save_ml_tokens():
    cfg["ml_access_token"]  = ml_access_token
    cfg["ml_refresh_token"] = ml_refresh_token
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


def refresh_ml_token():
    global ml_access_token, ml_refresh_token
    r = requests.post(
        "https://api.mercadolibre.com/oauth/token",
        data={
            "grant_type":    "refresh_token",
            "client_id":     ML_CLIENT_ID,
            "client_secret": ML_CLIENT_SECRET,
            "refresh_token": ml_refresh_token,
        },
        timeout=10,
    )
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError(f"ML token refresh failed: {d}")
    ml_access_token  = d["access_token"]
    ml_refresh_token = d.get("refresh_token", ml_refresh_token)
    save_ml_tokens()
    print("  [ML token refreshed]")


def ml_get(url, params=None, retry=True):
    """GET wrapper: auto-refresh on 401, back-off on 429."""
    global ml_access_token
    p = params or {}
    p["access_token"] = ml_access_token
    r = requests.get(url, params=p, timeout=10)
    if r.status_code == 401 and retry:
        refresh_ml_token()
        return ml_get(url, params, retry=False)
    if r.status_code == 429:
        print("  [ML rate limit – waiting 3 s]")
        time.sleep(3)
        return ml_get(url, params, retry=retry)
    return r


def ml_search_sku(sku):
    """Return list of MLB item IDs for this seller SKU."""
    r = ml_get(
        f"https://api.mercadolibre.com/users/{ML_USER_ID}/items/search",
        {"seller_sku": sku},
    )
    return r.json().get("results", [])


def ml_get_items_status(ids):
    """Return list of {id, status, migrated, item_group_id, catalog_product_id} for up to 20 item IDs."""
    if not ids:
        return []
    r = ml_get(
        "https://api.mercadolibre.com/items",
        {"ids": ",".join(ids[:20]), "attributes": "id,status,tags,item_group_id,catalog_product_id"},
    )
    out = []
    for entry in r.json():
        body = entry.get("body") or entry
        tags = body.get("tags") or []
        migrated = "variations_migration_source" in tags
        out.append({
            "id":                body.get("id", ""),
            "status":            body.get("status", ""),
            "migrated":          migrated,
            "item_group_id":     body.get("item_group_id", ""),
            "catalog_product_id": body.get("catalog_product_id", ""),
        })
    return out


def ml_get_user_product(item_id):
    """Fetch User Product (UP) for a given MLB item ID.
    Returns dict with family_id, catalog_product_id, up_id, up_status, up_item_id, or None if not found.

    Endpoint: GET /users/{user_id}/user_products?item_id={item_id}
    """
    r = ml_get(
        f"https://api.mercadolibre.com/users/{ML_USER_ID}/user_products",
        {"item_id": item_id},
    )
    data = r.json()
    # Response may be a list or {"results": [...]}
    results = data if isinstance(data, list) else data.get("results", [])
    if not results:
        return None
    up = results[0]
    return {
        "family_id":          up.get("family_id", ""),
        "catalog_product_id": up.get("catalog_product_id", ""),
        "up_id":              str(up.get("id", "")),
        "up_status":          up.get("status", ""),
        "up_item_id":         up.get("item_id", ""),
    }


def ml_find_active_for_migrated(old_item_id):
    """For an item with variations_migration_source tag, find the new MLB items via migration endpoint.
    Endpoint: GET /items/{id}/migration_live_listing
    Returns (first_new_id, summary_str) — both empty string if not found."""
    r = ml_get(f"https://api.mercadolibre.com/items/{old_item_id}/migration_live_listing")
    data = r.json()
    new_items = data.get("new_items", [])
    new_ids = [
        entry["new_item_id"]
        for entry in new_items
        if entry.get("new_item_id") and entry.get("migration_status") == "created"
    ]
    if not new_ids:
        # include pending too if no created ones yet
        new_ids = [entry["new_item_id"] for entry in new_items if entry.get("new_item_id")]
    if not new_ids:
        return "", ""
    unique_ids = list(dict.fromkeys(new_ids))  # deduplicate preserving order
    summary = ", ".join(unique_ids[:3])
    if len(unique_ids) > 3:
        summary += f" (+{len(unique_ids) - 3})"
    return unique_ids[0], summary


def is_family_id(val) -> bool:
    """Retorna True se val for um número puro com 14+ dígitos (family_id do ML)."""
    s = str(val or "").strip()
    return s.isdigit() and len(s) >= 14


def build_family_map() -> dict:
    """
    Pagina todos os itens do vendedor e mapeia family_id → [MLB IDs].
    A API ML ignora o filtro family_id no endpoint de busca; a única forma
    confiável é buscar o campo item a item em lotes de 20.
    """
    print("  Mapeando family_ids dos itens do vendedor (necessario para grupos)...")
    all_ids = []
    offset  = 0
    limit   = 50
    while True:
        r       = ml_get(f"https://api.mercadolibre.com/users/{ML_USER_ID}/items/search",
                         {"limit": limit, "offset": offset})
        data    = r.json()
        results = data.get("results", [])
        all_ids.extend(results)
        total   = data.get("paging", {}).get("total", 0)
        offset += limit
        if offset >= total or not results:
            break
        time.sleep(0.1)
    print(f"  Total de itens do vendedor: {len(all_ids)}")

    family_map = {}
    for i in range(0, len(all_ids), 20):
        chunk = all_ids[i:i+20]
        r     = ml_get("https://api.mercadolibre.com/items",
                       {"ids": ",".join(chunk), "attributes": "id,family_id,status"})
        data  = r.json()
        if isinstance(data, list):
            for entry in data:
                body = entry.get("body", {}) if isinstance(entry, dict) else {}
                fid  = str(body.get("family_id") or "").strip()
                mlb  = body.get("id", "")
                if fid and mlb:
                    family_map.setdefault(fid, []).append(mlb)
        time.sleep(0.1)

    families = len(family_map)
    mapped   = sum(len(v) for v in family_map.values())
    print(f"  Familias encontradas: {families} | Itens com family_id: {mapped}")
    return family_map


def ml_get_active_in_group(item_group_id):
    """Given an item_group_id, return list of active MLB IDs in that group."""
    if not item_group_id:
        return []
    r = ml_get(
        f"https://api.mercadolibre.com/users/{ML_USER_ID}/items/search",
        {"item_group_id": item_group_id},
    )
    return r.json().get("results", [])


def check_sku_ml(sku):
    """Check one SKU on ML. Returns list of {id, status, migrated}.
    If sku looks like an MLB ID, fetch it directly instead of searching by seller_sku."""
    if str(sku).upper().startswith("MLB"):
        items = ml_get_items_status([sku])
        time.sleep(0.15)
        return items
    ids = ml_search_sku(sku)
    time.sleep(0.15)
    if not ids:
        return []
    items = ml_get_items_status(ids)
    time.sleep(0.15)
    return items


# ── Tiny helpers (v3) ─────────────────────────────────────────────────────────

def tiny_search_sku(sku):
    """Search Tiny v3 by SKU (codigo field). Returns matching product dict or None.
    Skips search for MLB/MLBU codes — those are ML identifiers, not Tiny SKUs."""
    if str(sku).upper().startswith("MLB"):
        return None
    sku_str = str(sku).strip()
    try:
        result = tc.get("/produtos", {"codigo": sku_str})
        for item in result.get("itens", []):
            if str(item.get("sku", "")).strip() == sku_str:
                return item
    except Exception as e:
        print(f"  [Tiny error] {e}")
    return None


def tiny_get_product(product_id):
    """Fetch full product details (includes variacoes for kits)."""
    try:
        return tc.get(f"/produtos/{product_id}")
    except Exception as e:
        print(f"  [Tiny error] {e}")
        return None


# ── Column E logic ────────────────────────────────────────────────────────────

def check_kit_variations(variacoes):
    """
    Given a Tiny variacoes list, check each variation SKU on ML.
    Returns (col_e_kit_suffix, color) where suffix is like:
      'kit - variações ativas: MLB... (SKU ...) | ...'
    """
    active_parts = []
    paused_parts = []
    closed_parts = []

    for entry in variacoes:
        # Tiny v3: variation is a flat dict with "sku" field directly
        var_sku = str(entry.get("sku", "")).strip()
        if not var_sku:
            continue

        var_items = check_sku_ml(var_sku)

        for item in var_items:
            label = f"{item['id']} (SKU {var_sku})"
            if item["status"] == "active":
                active_parts.append(label)
            elif item["status"] == "paused":
                paused_parts.append(label)
            elif item["status"] == "closed":
                closed_parts.append(label)

    if not active_parts and not paused_parts and not closed_parts:
        return "kit - variações não encontradas no ML", "dark_red"

    parts = []
    if active_parts:
        parts.append("variações ativas: " + " | ".join(active_parts))
    if paused_parts:
        parts.append("variações inativas (paused): " + " | ".join(paused_parts))
    if closed_parts:
        parts.append("variações fechadas: " + " | ".join(closed_parts))

    return "kit - " + " | ".join(parts), "orange" if active_parts else "dark_red"


def build_col_e(sku):
    """
    Returns (col_e_value, color)
      color: "green" | "gray" | "red" | "yellow" | "purple" | "blue" | "orange" | "dark_red" | None

    Colors:
      green    = anúncio ativo no ML
      gray     = anúncio inativo (paused) no ML
      red      = anúncio fechado no ML
      yellow   = cadastrado no Tiny, sem anúncio no ML
      purple   = não encontrado em nenhum sistema
      blue     = migração / family encontrada
      orange   = kit com variações ativas no ML
      dark_red = kit sem variações no ML / variações fechadas
    """
    sku_str = str(sku).strip()

    # ── Family ID path (número puro 14+ dígitos) ─────────────────────────────
    if is_family_id(sku_str):
        members = _family_map.get(sku_str, [])
        if members:
            summary = ", ".join(members[:5])
            if len(members) > 5:
                summary += f" (+{len(members) - 5})"
            return f"Migração encontrada: {sku_str} → {len(members)} variações: {summary}", "blue"
        return f"Family ID {sku_str} — não encontrado", "purple"

    # ── MLB ID path ──────────────────────────────────────────────────────────
    if sku_str.upper().startswith("MLB"):
        ml_items = ml_get_items_status([sku_str])
        if not ml_items:
            return "não encontrado no ML nem no Tiny", "purple"
        item = ml_items[0]
        if item.get("migrated"):
            group_id = item.get("item_group_id", "")
            active_in_group = ml_get_active_in_group(group_id) if group_id else []
            time.sleep(0.15)
            if not active_in_group:
                _, up_summary = ml_find_active_for_migrated(item["id"])
                time.sleep(0.15)
                if up_summary:
                    return f"migração encontrada: {item['id']} → novos itens: {up_summary}", "blue"
            else:
                return f"migração encontrada: {item['id']} → grupo ativo: {', '.join(active_in_group)}", "blue"
            return f"migração encontrada: {item['id']}", "blue"
        status = item.get("status", "")
        # Fetch User Product to get family_id
        up = ml_get_user_product(sku_str)
        time.sleep(0.15)
        family_suffix = ""
        if up and up.get("family_id"):
            family_suffix = f" | family: {up['family_id']}"
        if status == "active":
            return f"anuncio ativo: {item['id']}{family_suffix}", "green"
        elif status == "paused":
            return f"anuncio inativo (paused): {item['id']}{family_suffix}", "gray"
        elif status == "closed":
            return f"anuncio fechado: {item['id']}{family_suffix}", "red"
        return f"anuncio {status}: {item['id']}{family_suffix}", None

    # ── SKU path ─────────────────────────────────────────────────────────────
    # 1 — Check ML
    ml_items = check_sku_ml(sku_str)

    ml_label = None
    ml_color = None
    if ml_items:
        active  = [i for i in ml_items if i["status"] == "active"]
        paused  = [i for i in ml_items if i["status"] == "paused"]
        closed  = [i for i in ml_items if i["status"] == "closed"]
        if active:
            up = ml_get_user_product(active[0]["id"])
            time.sleep(0.15)
            family_suffix = f" | family: {up['family_id']}" if up and up.get("family_id") else ""
            ml_label = f"anuncio ativo: {', '.join(i['id'] for i in active)}{family_suffix}"
            ml_color = "green"
        elif paused:
            up = ml_get_user_product(paused[0]["id"])
            time.sleep(0.15)
            family_suffix = f" | family: {up['family_id']}" if up and up.get("family_id") else ""
            ml_label = f"anuncio inativo (paused): {', '.join(i['id'] for i in paused)}{family_suffix}"
            ml_color = "gray"
        elif closed:
            if any(i.get("migrated") for i in closed):
                closed_ids = ', '.join(i['id'] for i in closed if i.get("migrated"))
                group_id = next((i.get("item_group_id") for i in closed if i.get("migrated") and i.get("item_group_id")), None)
                active_in_group = []
                if group_id:
                    active_in_group = ml_get_active_in_group(group_id)
                    time.sleep(0.15)
                if not active_in_group:
                    migrated_id = next((i["id"] for i in closed if i.get("migrated")), None)
                    if migrated_id:
                        _, up_summary = ml_find_active_for_migrated(migrated_id)
                        time.sleep(0.15)
                        if up_summary:
                            ml_label = f"migração encontrada: {closed_ids} → novos itens: {up_summary}"
                        else:
                            ml_label = f"migração encontrada: {closed_ids}"
                    else:
                        ml_label = f"migração encontrada: {closed_ids}"
                else:
                    ml_label = f"migração encontrada: {closed_ids} → grupo ativo: {', '.join(active_in_group)}"
                ml_color = "blue"
            else:
                ml_label = f"anuncio fechado: {', '.join(i['id'] for i in closed)}"
                ml_color = "red"

    # 2 — Check Tiny (always, even if found on ML)
    tiny_prod = tiny_search_sku(sku_str)
    time.sleep(0.2)

    if not tiny_prod:
        if ml_label:
            return ml_label, ml_color
        return "não encontrado no ML nem no Tiny", "purple"

    product_id = tiny_prod.get("id")
    full_prod   = tiny_get_product(product_id)
    time.sleep(0.2)

    if not full_prod:
        if ml_label:
            return ml_label, ml_color
        return "cadastrado no Tiny | sem anúncio no ML", "yellow"

    variacoes = full_prod.get("variacoes", [])

    # 3 — Simple product in Tiny (no variations/kit)
    if not variacoes:
        if ml_label:
            return ml_label, ml_color
        tiny_status = "ativo" if full_prod.get("situacao") == "A" else full_prod.get("situacao", "?")
        return f"cadastrado no Tiny ({tiny_status}) | sem anúncio no ML", "yellow"

    # 4 — Kit in Tiny: check each variation on ML
    kit_suffix, kit_color = check_kit_variations(variacoes)

    if ml_label:
        combined = f"{ml_label} | {kit_suffix}"
        return combined, kit_color or ml_color
    else:
        return kit_suffix, kit_color


# ── Main ──────────────────────────────────────────────────────────────────────

GREEN_FILL    = PatternFill("solid", fgColor="92D050")  # verde   — anúncio ativo
GRAY_FILL     = PatternFill("solid", fgColor="D9D9D9")  # cinza   — anúncio paused
RED_FILL      = PatternFill("solid", fgColor="FF6B6B")  # vermelho — anúncio fechado
YELLOW_FILL   = PatternFill("solid", fgColor="FFFF99")  # amarelo  — só no Tiny
PURPLE_FILL   = PatternFill("solid", fgColor="D9B3FF")  # roxo     — não encontrado
BLUE_FILL     = PatternFill("solid", fgColor="9DC3E6")  # azul     — migração
ORANGE_FILL   = PatternFill("solid", fgColor="FFB347")  # laranja  — kit ativo
DARK_RED_FILL = PatternFill("solid", fgColor="C00000")  # verm.esc — kit sem ML
NO_FILL       = PatternFill("none")

# Rows with these values will be re-checked (not skipped)
RECHECK_VALUES = {
    "não encontrado no ML",
    "não encontrado no ML nem no Tiny",
    "cadastrado no Tiny | sem anúncio no ML",
}

def should_recheck(col_e_val):
    """Return True if this row should be reprocessed."""
    if not col_e_val:
        return True
    if col_e_val in RECHECK_VALUES:
        return True
    if col_e_val.startswith("kit -"):
        return True
    return False


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-row",  type=int, default=0)
    parser.add_argument("--to-row",    type=int, default=0)
    parser.add_argument("--force-all", action="store_true",
                        help="Reprocessa todas as linhas, sobrescrevendo valores existentes")
    args = parser.parse_args()

    # Refresh ML token at startup
    print("Refreshing ML token...")
    try:
        refresh_ml_token()
    except Exception as e:
        print(f"WARNING: {e}")
        if not ml_access_token:
            sys.exit("No valid ML token. Aborting.")

    print(f"Loading: {EXCEL_INPUT}")
    wb = load_workbook(EXCEL_INPUT)
    ws = wb.active

    # Pré-scan: verifica se alguma linha tem family_id para construir o mapa
    has_family = any(
        is_family_id(row[3].value)
        for row in ws.iter_rows(min_row=2)
        if row[3].value is not None
    )
    if has_family:
        _family_map.update(build_family_map())

    total   = 0
    updated = 0
    skipped = 0
    errors  = 0

    for row in ws.iter_rows(min_row=2):
        raw_d = row[3].value
        if isinstance(raw_d, float) and raw_d == int(raw_d):
            raw_d = int(raw_d)
        col_d = str(raw_d or "").strip()
        col_e = str(row[4].value or "").strip()
        row_n = row[0].row

        if not col_d:
            continue

        total += 1

        # Hard stop: don't process anything beyond --to-row
        if args.to_row and row_n > args.to_row:
            break

        in_forced_range = args.from_row and args.to_row and args.from_row <= row_n <= args.to_row

        # Skip already-filled rows (unless forced)
        if not args.force_all and not in_forced_range and not should_recheck(col_e):
            skipped += 1
            continue

        print(f"Row {row_n}: {col_d} ...", end=" ", flush=True)

        try:
            new_val, color = build_col_e(col_d)

            ws.cell(row=row_n, column=5).value = new_val

            if color == "green":
                fill = GREEN_FILL
            elif color == "gray":
                fill = GRAY_FILL
            elif color == "red":
                fill = RED_FILL
            elif color == "yellow":
                fill = YELLOW_FILL
            elif color == "purple":
                fill = PURPLE_FILL
            elif color == "blue":
                fill = BLUE_FILL
            elif color == "orange":
                fill = ORANGE_FILL
            elif color == "dark_red":
                fill = DARK_RED_FILL
            else:
                fill = NO_FILL
            for col in range(1, 6):
                ws.cell(row=row_n, column=col).fill = fill

            print(f"→ {new_val[:70]}")
            updated += 1

            # Checkpoint every 50 updates
            if updated % 50 == 0:
                wb.save(EXCEL_OUTPUT)
                print(f"  [Checkpoint: {updated} rows saved]")

        except Exception as e:
            print(f"ERROR: {e}")
            errors += 1

    wb.save(EXCEL_OUTPUT)

    print("\n" + "=" * 60)
    print(f"Done.")
    print(f"  Total rows with SKU : {total}")
    print(f"  Updated             : {updated}")
    print(f"  Skipped (pre-filled): {skipped}")
    print(f"  Errors              : {errors}")
    print(f"  Output saved to     : {EXCEL_OUTPUT}")


if __name__ == "__main__":
    main()
